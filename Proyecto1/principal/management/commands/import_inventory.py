from django.core.management.base import BaseCommand
from principal.models import Lot
from principal.views import normalize_lot_id  # o mueve normalize_lot_id a utils.py
import openpyxl
from pathlib import Path
from django.conf import settings

class Command(BaseCommand):
    help = "Importa inventory.xlsx a la tabla Lot (una sola vez)."

    def handle(self, *args, **options):
        path = Path(settings.BASE_DIR) / "principal" / "data" / "inventory.xlsx"
        if not path.exists():
            self.stdout.write(self.style.ERROR(f"No existe: {path}"))
            return

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]

        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        headers = [str(h).strip() if h else "" for h in headers]

        created = 0
        updated = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            d = dict(zip(headers, row))
            raw_id = str(d.get("id_lote", "")).strip().upper()
            lot_id = normalize_lot_id(raw_id) or raw_id
            if not lot_id:
                continue

            obj, is_created = Lot.objects.update_or_create(
                id_lote=lot_id,
                defaults={
                    "estado_lote": d.get("estado_lote"),
                    "precio_total": d.get("precio_total"),
                    "precio_m2": d.get("precio_m2"),
                    "superficie_m2": d.get("superficie_m2"),
                    "proyecto": d.get("proyecto"),
                    "manzana": d.get("manzana"),
                    "medidas_lotes": d.get("medidas_lotes"),
                    "cantidad_de_apartado": d.get("cantidad_de_apartado"),
                    "dias_limite_apartado": d.get("dias_limite_apartado"),
                    "cantidad_enganche": d.get("cantidad_enganche"),
                    "cantidad_financiamiento": d.get("cantidad_financiamiento"),
                    "pago_mensualidad": d.get("pago_mensualidad"),
                    "cantidad_liquidacion": d.get("cantidad_liquidacion"),
                    "url_imagen_lote": d.get("url_imagen_lote"),
                },
            )
            if is_created:
                created += 1
            else:
                updated += 1

        self.stdout.write(self.style.SUCCESS(f"Listo. created={created}, updated={updated}"))
