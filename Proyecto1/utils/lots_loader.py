from __future__ import annotations
import csv, io, urllib.request
from typing import Dict, Any
from django.conf import settings

# ---------------------------
# Normalización de ESTADOS
# ---------------------------
VALID = {"disponible", "vendido", "apartado"}
NORMALIZE = {
    "vendible": "vendido",
    "vendido": "vendido",
    "disponible": "disponible",
    "apartado": "apartado",
}

# Columns esperadas (case-insensitive)
REQUIRED_MIN = {"development", "number", "status"}
OPTIONAL_COLS = {
    # Lote
    "area_m2",
    "price_total",  # precio de lista
    "precio_m2",
    "regular_width_m",
    "regular_length_m",
    "irregular_sides_m",  # "10,30,9.8,29.7"
    "image_filename",
    # Etapa (centralizado pero administrable)
    "stage_name",
    "stage_down_payment_percent",  # % enganche (default 20)
    "stage_apartar_amount",        # apartado (default 1000)
    "stage_deadline_days",         # límite días (default 5)
}

# Caches (separados para compatibilidad)
_CACHE_STATES: Dict[str, Dict[int, str]] | None = None
_CACHE_INVENTORY: Dict[str, Dict[int, Dict[str, Any]]] | None = None


# ---------- util I/O ----------
def _fetch_bytes(url_or_path: str, timeout: int = 15) -> bytes:
    if url_or_path.startswith(("http://", "https://")):
        with urllib.request.urlopen(url_or_path, timeout=timeout) as r:
            return r.read()
    with open(url_or_path, "rb") as f:
        return f.read()


# ---------- helpers parsing ----------
def _to_float(v: Any, default: float = 0.0) -> float:
    if v in ("", None):
        return default
    try:
        return float(str(v).replace(",", "").strip())
    except Exception:
        return default


def _to_int(v: Any, default: int | None = None) -> int | None:
    try:
        return int(str(v).strip())
    except Exception:
        return default


def _parse_csv_inventory(data: bytes) -> Dict[str, Dict[int, Dict[str, Any]]]:
    """
    Devuelve:
    {
      dev_slug: {
        number: {
          "status": "...",
          "area_m2": 300.0,
          "precio_m2: 900.0,
          "price_total": 1530000.0,
          "regular_width_m": 10.0 | None,
          "regular_length_m": 30.0 | None,
          "irregular_sides_m": "10,30,9.8,29.7",
          "image_filename": "lote-24.png",
          "stage_name": "Etapa 1",
          "stage_down_payment_percent": 20.0,
          "stage_apartar_amount": 1000.0,
          "stage_deadline_days": 5,
        }, ...
      }
    }
    """
    out: Dict[str, Dict[int, Dict[str, Any]]] = {}
    text = data.decode("utf-8-sig").splitlines()
    reader = csv.DictReader(text)
    # normalizar headers
    headers = [h.strip().lower() for h in reader.fieldnames or []]

    def g(row: dict, key: str, default: str = "") -> str:
        # acceso case-insensitive
        for i, h in enumerate(headers):
            if h == key:
                return str(list(row.values())[i] if i < len(row) else default)
        return default

    for row in reader:
        dev = g(row, "development").strip().lower()
        num_s = g(row, "number").strip()
        st = g(row, "status").strip().lower()
        st = NORMALIZE.get(st, st)

        if not dev or st not in VALID:
            continue

        n = _to_int(num_s)
        if n is None:
            continue

        item: Dict[str, Any] = {
            "status": st,
            "area_m2": _to_float(g(row, "area_m2")),
            "precio_m2": _to_float(g(row, "precio_m2")),
            "price_total": _to_float(g(row, "price_total")),
            "regular_width_m": (_to_float(g(row, "regular_width_m")) if g(row, "regular_width_m").strip() else None),
            "regular_length_m": (_to_float(g(row, "regular_length_m")) if g(row, "regular_length_m").strip() else None),
            "irregular_sides_m": g(row, "irregular_sides_m").strip(),
            "image_filename": g(row, "image_filename").strip(),
            "stage_name": g(row, "stage_name").strip(),
            "stage_down_payment_percent": _to_float(g(row, "stage_down_payment_percent"), 20.0) or 20.0,
            "stage_apartar_amount": _to_float(g(row, "stage_apartar_amount"), 1000.0) or 1000.0,
            "stage_deadline_days": _to_int(g(row, "stage_deadline_days"), 5) or 5,
        }
        out.setdefault(dev, {})[n] = item
    return out


def _parse_xlsx_inventory(data: bytes) -> Dict[str, Dict[int, Dict[str, Any]]]:
    from openpyxl import load_workbook

    out: Dict[str, Dict[int, Dict[str, Any]]] = {}
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    # Hoja por defecto: si existe "estados" o la primera con headers válidos
    ws = wb["estados"] if "estados" in wb.sheetnames else None
    if ws is None:
        for name in wb.sheetnames:
            _ws = wb[name]
            row1 = next(_ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not row1:
                continue
            header = [(str(c).strip().lower() if c is not None else "") for c in row1]
            if REQUIRED_MIN.issubset(set(header)):
                ws = _ws
                break
    if ws is None:
        return out

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return out
    header = [(str(c).strip().lower() if c is not None else "") for c in header_row]
    if not REQUIRED_MIN.issubset(set(header)):
        return out

    # índices case-insensitive
    idx = {k: header.index(k) for k in REQUIRED_MIN if k in header}

    # helpers para obtener por nombre
    def get_val(r: tuple, key: str) -> Any:
        if key in idx:
            i = idx[key]
            return r[i] if i < len(r) else ""
        # buscar si es opcional
        if key in OPTIONAL_COLS and key in header:
            j = header.index(key)
            return r[j] if j < len(r) else ""
        return ""

    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r:
            continue

        dev = str(get_val(r, "development") or "").strip().lower()
        num_raw = get_val(r, "number")
        st = str(get_val(r, "status") or "").strip().lower()
        st = NORMALIZE.get(st, st)
        if not dev or st not in VALID:
            continue

        n = _to_int(num_raw)
        if n is None:
            continue

        item: Dict[str, Any] = {
            "status": st,
            "area_m2": _to_float(get_val(r, "area_m2")),
            "precio_m2": _to_float(get_val(r, "precio_m2")),
            "price_total": _to_float(get_val(r, "price_total")),
            "regular_width_m": (_to_float(get_val(r, "regular_width_m")) if str(get_val(r, "regular_width_m")).strip() else None),
            "regular_length_m": (_to_float(get_val(r, "regular_length_m")) if str(get_val(r, "regular_length_m")).strip() else None),
            "irregular_sides_m": str(get_val(r, "irregular_sides_m") or "").strip(),
            "image_filename": str(get_val(r, "image_filename") or "").strip(),
            "stage_name": str(get_val(r, "stage_name") or "").strip(),
            "stage_down_payment_percent": _to_float(get_val(r, "stage_down_payment_percent"), 20.0) or 20.0,
            "stage_apartar_amount": _to_float(get_val(r, "stage_apartar_amount"), 1000.0) or 1000.0,
            "stage_deadline_days": _to_int(get_val(r, "stage_deadline_days"), 5) or 5,
        }
        out.setdefault(dev, {})[n] = item
    return out


# ---------- APIs públicas ----------
def load_inventory(force_reload: bool = False) -> Dict[str, Dict[int, Dict[str, Any]]]:
    """
    Lee el archivo (CSV/XLSX) y regresa un inventario rico:
    { dev: { number: {status, area_m2, price_total, medidas..., stage_*...} } }
    """
    global _CACHE_INVENTORY
    if _CACHE_INVENTORY is not None and not force_reload:
        return _CACHE_INVENTORY

    url = getattr(settings, "LOTS_SHEET_URL", "") or ""
    fmt = (getattr(settings, "LOTS_SHEET_FORMAT", "xlsx") or "xlsx").lower()
    timeout = int(getattr(settings, "LOTS_SHEET_TIMEOUT", 15))

    if not url:
        _CACHE_INVENTORY = {}
        return _CACHE_INVENTORY

    data = _fetch_bytes(url, timeout=timeout)
    _CACHE_INVENTORY = _parse_csv_inventory(data) if fmt == "csv" else _parse_xlsx_inventory(data)
    return _CACHE_INVENTORY


def load_states(force_reload: bool = False) -> Dict[str, Dict[int, str]]:
    """
    Mantiene compatibilidad con tu código anterior:
    Devuelve {development_slug: {lote_num: status}}
    a partir del inventario completo.
    """
    global _CACHE_STATES
    if _CACHE_STATES is not None and not force_reload:
        return _CACHE_STATES

    inv = load_inventory(force_reload=force_reload)
    states: Dict[str, Dict[int, str]] = {}
    for dev, lots in inv.items():
        for num, info in lots.items():
            states.setdefault(dev, {})[num] = str(info.get("status", "disponible"))
    _CACHE_STATES = states
    return _CACHE_STATES
